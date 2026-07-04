from decimal import Decimal
from io import BytesIO
import re
from threading import Lock, Thread
from uuid import uuid4

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.exc import IntegrityError

from app.core.audit_types import ENTITY_TYPE_PRODUCT, EVENT_TYPE_PRODUCT_CREATE, EVENT_TYPE_PRODUCT_DELETE, EVENT_TYPE_PRODUCT_UPDATE
from app.models.reference_data import Product
from app.repositories.reference_data import ProductRepository, ReferenceDataRepository
from app.schemas.common import build_pagination, model_to_dict
from app.schemas.reference_data import CurrencyPayload, EstablishmentPayload, OrderMethodPayload, ProductCreatePayload, ProductUpdatePayload, StatusPayload
from app.services.audit import log_audit_event
from app.services.price_federation import hydrate_products_async, hydrate_products_for_names
from app.services.serializers import serialize_currency, serialize_establishment, serialize_order_method, serialize_product, serialize_status


DEFAULT_ESTABLISHMENTS = [
    {"establishment_name": "Белка", "establishment_address": None},
    {"establishment_name": "Окто", "establishment_address": None},
    {"establishment_name": "Тула", "establishment_address": None},
]
DEFAULT_ORDER_METHODS = [
    {"order_method_name": "Курьер"},
    {"order_method_name": "Самовывоз"},
    {"order_method_name": "СДЭК"},
    {"order_method_name": "Яндекс"},
    {"order_method_name": "5Post"},
    {"order_method_name": "Авито", "order_method_sub_methods": ["Авито", "СДЭК", "Яндекс", "5Post", "Почта"]},
]
DEFAULT_STATUSES = [
    {"status_type": "orders", "status_status": "Новый", "status_color": "orange"},
    {"status_type": "orders", "status_status": "В обработке", "status_color": "#3b82f6"},
    {"status_type": "orders", "status_status": "На сборку", "status_color": "#6366f1"},
    {"status_type": "orders", "status_status": "Собран", "status_color": "#0f766e"},
    {"status_type": "orders", "status_status": "Выполнен", "status_color": "#16a34a"},
    {"status_type": "orders", "status_status": "Отменен", "status_color": "red"},
    {"status_type": "orders", "status_status": "Возврат", "status_color": "#9a8b2f"},
    {"status_type": "inventory", "status_status": "Новый", "status_color": "orange"},
    {"status_type": "inventory", "status_status": "В обработке", "status_color": "blue"},
    {"status_type": "inventory", "status_status": "Завершена", "status_color": "green"},
    {"status_type": "order_products", "status_status": "Не обработан", "status_color": "gray"},
    {"status_type": "order_products", "status_status": "Перемещение", "status_color": "blue"},
    {"status_type": "order_products", "status_status": "Заказ поставщику", "status_color": "orange"},
    {"status_type": "order_products", "status_status": "Собрано", "status_color": "#8b5cf6"},
    {"status_type": "order_products", "status_status": "Не будет", "status_color": "red"},
    {"status_type": "order_products", "status_status": "В наличии", "status_color": "green"},
    {"status_type": "order_products", "status_status": "Отгружено", "status_color": "blue"},
    {"status_type": "order_products", "status_status": "Отменен", "status_color": "red"},
    {"status_type": "order_products", "status_status": "Возврат", "status_color": "#9a8b2f"},
    {"status_type": "product_registration", "status_status": "Новый", "status_color": "orange"},
    {"status_type": "product_registration", "status_status": "В обработке", "status_color": "blue"},
    {"status_type": "product_registration", "status_status": "Принято на складе", "status_color": "green"},
]

OBSOLETE_DEFAULT_STATUSES = {
    ("order_products", "Принято на складе"),
}
DEFAULT_CURRENCIES = [
    {"currency_name": "USD", "currency_sign": "$"},
    {"currency_name": "RUB", "currency_sign": "₽"},
]
DEFAULT_PRODUCTS = [
    {
        "product_article": "000009",
        "product_name": "Balenciaga: (10, Avenue George V) - Eau de Parfum",
        "product_cost_usd": Decimal("0.00"),
    }
]

_ZERO_PAD_FORMAT_RE = re.compile(r"^0+$")
_IMPORT_COST_QUANT = Decimal("0.01")


class ProductImportJobStore:
    def __init__(self) -> None:
        self._lock = Lock()
        self._jobs: dict[str, dict] = {}

    def create(self, *, filename: str | None, created_by_user_id: int) -> dict:
        job_id = uuid4().hex
        job = {
            "job_id": job_id,
            "filename": filename,
            "status": "queued",
            "created_by_user_id": created_by_user_id,
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "processed": 0,
            "total_rows": 0,
            "total_products": 0,
            "error_message": None,
        }
        with self._lock:
            self._jobs[job_id] = job
        return dict(job)

    def get(self, job_id: str) -> dict | None:
        with self._lock:
            job = self._jobs.get(job_id)
            return dict(job) if job is not None else None

    def update(self, job_id: str, **changes) -> dict:
        with self._lock:
            job = self._jobs.get(job_id)
            if job is None:
                raise KeyError(job_id)
            job.update(changes)
            return dict(job)


_product_import_jobs = ProductImportJobStore()
_reference_seed_lock = Lock()


def _normalize_import_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text) or None


def _normalize_article_cell(cell) -> str | None:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, str):
        return _normalize_import_text(value)
    if isinstance(value, (int, Decimal)):
        integer_value = int(value)
        number_format = (cell.number_format or "").strip()
        if _ZERO_PAD_FORMAT_RE.fullmatch(number_format):
            return str(integer_value).zfill(len(number_format))
        return str(integer_value)
    if isinstance(value, float):
        if value.is_integer():
            integer_value = int(value)
            number_format = (cell.number_format or "").strip()
            if _ZERO_PAD_FORMAT_RE.fullmatch(number_format):
                return str(integer_value).zfill(len(number_format))
            return str(integer_value)
        return _normalize_import_text(value)
    return _normalize_import_text(value)


def _parse_cost_cell(cell) -> Decimal | None:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value.quantize(_IMPORT_COST_QUANT)
    if isinstance(value, int):
        return Decimal(value).quantize(_IMPORT_COST_QUANT)
    if isinstance(value, float):
        return Decimal(str(value)).quantize(_IMPORT_COST_QUANT)

    text = _normalize_import_text(value)
    if text is None:
        return None
    normalized = text.replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized).quantize(_IMPORT_COST_QUANT)
    except Exception:
        return None


def _is_header_row(article: str | None, name: str | None, cost_cell_value: object) -> bool:
    article_value = (article or "").strip().lower()
    name_value = (name or "").strip().lower()
    cost_value = (_normalize_import_text(cost_cell_value) or "").strip().lower()
    return article_value in {"article", "артикул"} and name_value in {"name", "наименование", "товар"} and cost_value in {"cost", "cost_usd", "цена", "цена usd", "usd"}


def _iter_products_from_xlsx(content: bytes) -> list[dict]:
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Файл пустой")

    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Не удалось прочитать XLSX файл") from exc

    try:
        sheet = workbook.active
        items: list[dict] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=1), start=1):
            article_cell = row[0] if len(row) >= 1 else None
            name_cell = row[1] if len(row) >= 2 else None
            cost_cell = row[3] if len(row) >= 4 else None

            article = _normalize_article_cell(article_cell) if article_cell is not None else None
            name = _normalize_import_text(name_cell.value) if name_cell is not None else None
            cost = _parse_cost_cell(cost_cell) if cost_cell is not None else None

            if _is_header_row(article, name, cost_cell.value if cost_cell is not None else None):
                continue

            items.append(
                {
                    "row_index": row_index,
                    "product_article": article,
                    "product_name": name,
                    "product_cost_usd": cost,
                }
            )
        return items
    finally:
        workbook.close()


def _run_product_import_job(*, job_id: str, parsed_rows: list[dict], current_user: dict, session_factory) -> None:
    db = session_factory()
    try:
        repository = ProductRepository(db)
        created = 0
        updated = 0
        skipped = 0
        processed = 0

        _product_import_jobs.update(job_id, status="running", total_rows=len(parsed_rows), processed=0, created=0, updated=0, skipped=0, error_message=None)

        for item in parsed_rows:
            article = item["product_article"]
            name = item["product_name"]
            cost = item["product_cost_usd"]

            if article is None or name is None or cost is None:
                skipped += 1
                processed += 1
                if processed % 100 == 0 or processed == len(parsed_rows):
                    _product_import_jobs.update(job_id, processed=processed, created=created, updated=updated, skipped=skipped)
                continue

            row = repository.get_by_article(article)
            if row is None:
                db.add(
                    Product(
                        product_article=article,
                        product_name=name,
                        product_cost_usd=cost,
                        product_owner_user_id=current_user["user_id"],
                    )
                )
                created += 1
            else:
                changed = False
                if row.product_name != name:
                    row.product_name = name
                    changed = True
                if row.product_cost_usd != cost:
                    row.product_cost_usd = cost
                    changed = True
                if row.product_owner_user_id != current_user["user_id"]:
                    row.product_owner_user_id = current_user["user_id"]
                    changed = True
                if changed:
                    updated += 1

            processed += 1
            if processed % 500 == 0:
                db.commit()
            if processed % 100 == 0 or processed == len(parsed_rows):
                _product_import_jobs.update(job_id, processed=processed, created=created, updated=updated, skipped=skipped)

        db.commit()
        total_products = db.query(Product).count()
        _product_import_jobs.update(
            job_id,
            status="completed",
            processed=processed,
            created=created,
            updated=updated,
            skipped=skipped,
            total_products=total_products,
        )
    except Exception as error:
        db.rollback()
        _product_import_jobs.update(job_id, status="failed", error_message=str(error))
    finally:
        db.close()


class ReferenceDataService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.repository = ReferenceDataRepository(db)
        self.product_repository = ProductRepository(db)

    def ensure_seed_data(self) -> None:
        with _reference_seed_lock:
            existing_establishments = {
                (item.establishment_name or "").strip(): item
                for item in self.repository.list_establishments()
            }
            for item in DEFAULT_ESTABLISHMENTS:
                establishment_name = (item["establishment_name"] or "").strip()
                existing_establishment = existing_establishments.get(establishment_name) or self.repository.get_establishment_by_name(establishment_name)
                if existing_establishment is None:
                    existing_establishment = self._create_seed_row_safely(
                        create_row=lambda: self.repository.create_establishment(item),
                        fetch_row=lambda: self.repository.get_establishment_by_name(establishment_name),
                    )
                if existing_establishment is not None:
                    existing_establishments[establishment_name] = existing_establishment

            for item in DEFAULT_ORDER_METHODS:
                order_method_name = item["order_method_name"]
                existing_method = self.repository.get_order_method_by_name(order_method_name)
                if existing_method is None:
                    existing_method = self._create_seed_row_safely(
                        create_row=lambda: self.repository.create_order_method(item),
                        fetch_row=lambda: self.repository.get_order_method_by_name(order_method_name),
                    )
                if existing_method is None:
                    continue
                expected_sub_methods = item.get("order_method_sub_methods")
                if existing_method.order_method_sub_methods != expected_sub_methods:
                    self.repository.update_row(existing_method, {"order_method_sub_methods": expected_sub_methods})

            existing_statuses = {
                ((item.status_type or "").strip(), (item.status_status or "").strip()): item
                for item in self.repository.list_statuses()
            }
            for item in DEFAULT_STATUSES:
                key = ((item["status_type"] or "").strip(), (item["status_status"] or "").strip())
                existing_status = existing_statuses.get(key) or self.repository.get_status_by_type_and_name(key[0], key[1])
                if existing_status is None:
                    existing_status = self._create_seed_row_safely(
                        create_row=lambda: self.repository.create_status(item),
                        fetch_row=lambda: self.repository.get_status_by_type_and_name(key[0], key[1]),
                    )
                if existing_status is None:
                    continue
                existing_statuses[key] = existing_status
                # Цвет дефолтных статусов задаётся только при создании (в create_status выше).
                # Существующие статусы не трогаем — иначе цвет, выставленный в админке,
                # затирался бы при каждом обращении к справочникам и не доезжал до приложения.

            for status_type, status_name in OBSOLETE_DEFAULT_STATUSES:
                obsolete_status = self.repository.get_status_by_type_and_name(status_type, status_name)
                if obsolete_status is not None:
                    self.db.delete(obsolete_status)
                    self.db.commit()

            existing_currencies = {
                (item.currency_name or "").strip().upper(): item
                for item in self.repository.list_currencies()
            }
            for item in DEFAULT_CURRENCIES:
                currency_name = (item["currency_name"] or "").strip().upper()
                existing_currency = existing_currencies.get(currency_name) or self.repository.get_currency_by_name(currency_name)
                if existing_currency is None:
                    existing_currency = self._create_seed_row_safely(
                        create_row=lambda: self.repository.create_currency(item),
                        fetch_row=lambda: self.repository.get_currency_by_name(currency_name),
                    )
                if existing_currency is not None:
                    existing_currencies[currency_name] = existing_currency

            for item in DEFAULT_PRODUCTS:
                product_article = item["product_article"]
                if self.product_repository.get_by_article(product_article) is None:
                    self._create_seed_row_safely(
                        create_row=lambda: self.product_repository.create(item),
                        fetch_row=lambda: self.product_repository.get_by_article(product_article),
                    )

    def _create_seed_row_safely(self, *, create_row, fetch_row):
        try:
            return create_row()
        except IntegrityError:
            self.db.rollback()
            return fetch_row()

    def get_reference_data(self) -> dict:
        self.ensure_seed_data()
        return {
            "establishments": [serialize_establishment(item) for item in self.repository.list_establishments()],
            "order_methods": [serialize_order_method(item) for item in self.repository.list_order_methods()],
            "statuses": [serialize_status(item) for item in self.repository.list_statuses()],
            "currencies": [serialize_currency(item) for item in self.repository.list_currencies()],
        }

    def list_establishments(self) -> dict:
        self.ensure_seed_data()
        return {"items": [serialize_establishment(item) for item in self.repository.list_establishments()]}

    def list_order_methods(self) -> dict:
        self.ensure_seed_data()
        return {"items": [serialize_order_method(item) for item in self.repository.list_order_methods()]}

    def list_statuses(self, *, status_type: str | None = None) -> dict:
        self.ensure_seed_data()
        return {"items": [serialize_status(item) for item in self.repository.list_statuses(status_type=status_type)]}

    def list_currencies(self) -> dict:
        self.ensure_seed_data()
        return {"items": [serialize_currency(item) for item in self.repository.list_currencies()]}

    def create_establishment(self, payload: EstablishmentPayload, current_user: dict) -> dict:
        row = self.repository.create_establishment({**model_to_dict(payload), "establishment_owner_user_id": current_user["user_id"]})
        return serialize_establishment(row)

    def create_order_method(self, payload: OrderMethodPayload, current_user: dict) -> dict:
        row = self.repository.create_order_method({**model_to_dict(payload), "order_method_owner_user_id": current_user["user_id"]})
        return serialize_order_method(row)

    def create_status(self, payload: StatusPayload, current_user: dict) -> dict:
        row = self.repository.create_status({**model_to_dict(payload), "status_owner_user_id": current_user["user_id"]})
        return serialize_status(row)

    def create_currency(self, payload: CurrencyPayload, current_user: dict) -> dict:
        row = self.repository.create_currency({**model_to_dict(payload), "currency_owner_user_id": current_user["user_id"]})
        return serialize_currency(row)

    def update_establishment(self, establishment_id: int, payload: EstablishmentPayload) -> dict:
        row = self.repository.get_establishment(establishment_id)
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Точка не найдена")
        return serialize_establishment(self.repository.update_row(row, model_to_dict(payload)))

    def update_order_method(self, order_method_id: int, payload: OrderMethodPayload) -> dict:
        row = self.repository.get_order_method(order_method_id)
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Способ заказа не найден")
        return serialize_order_method(self.repository.update_row(row, model_to_dict(payload)))

    def update_status(self, status_id: int, payload: StatusPayload) -> dict:
        row = self.repository.get_status(status_id)
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Статус не найден")
        return serialize_status(self.repository.update_row(row, model_to_dict(payload)))

    def update_currency(self, currency_id: int, payload: CurrencyPayload) -> dict:
        row = self.repository.get_currency(currency_id)
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Валюта не найдена")
        return serialize_currency(self.repository.update_row(row, model_to_dict(payload)))

    def list_products(self, *, search: str | None = None, page: int = 1, page_size: int = 20, sort_by: str = "product_id", sort_order: str = "desc") -> dict:
        self.ensure_seed_data()
        # Свежий каталог — из nn_vla (источник правды). Гидратация идёт в ФОНЕ: поиск
        # отвечает сразу локальными результатами (быстро), свежий CL подтягивается
        # отдельным потоком и появляется при следующем поиске.
        if search and search.strip():
            hydrate_products_async(search)
        rows, total = self.product_repository.list(search=search, page=page, page_size=page_size, sort_by=sort_by, sort_order=sort_order)
        return {
            "items": [serialize_product(item) for item in rows],
            "pagination": build_pagination(page, page_size, total),
            "filters": {"search": search, "sort_by": sort_by, "sort_order": sort_order},
        }

    def match_products_by_name(self, names: list[str]) -> dict:
        # Сопоставление списка наименований с номенклатурой (для импорта документа в заказ).
        # Точное совпадение по имени -> matched; несколько точных -> candidates;
        # ни одного -> подсказки по подстроке (candidates), matched=None.
        self.ensure_seed_data()
        # Свежая номенклатура — из nn_vla: до-заносим позиции под импортируемые имена,
        # чтобы матч по имени находил их (иначе «N позиций не найдено» из-за старого каталога).
        hydrate_products_for_names(self.db, names)
        results = []
        for raw in names:
            name = (raw or "").strip()
            if not name:
                results.append({"query": raw, "matched": None, "candidates": []})
                continue
            exact = self.product_repository.find_by_exact_name(name)
            if len(exact) == 1:
                results.append({"query": raw, "matched": serialize_product(exact[0]), "candidates": []})
            elif len(exact) > 1:
                results.append({"query": raw, "matched": None, "candidates": [serialize_product(p) for p in exact]})
            else:
                suggestions, _ = self.product_repository.list(search=name, page=1, page_size=5)
                results.append({"query": raw, "matched": None, "candidates": [serialize_product(p) for p in suggestions]})
        return {"results": results}

    def get_product_or_404(self, product_id: int):
        row = self.product_repository.get_by_id(product_id)
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Товар не найден")
        return row

    def _generate_unique_article(self) -> str:
        # Уникальный авто-артикул для товаров, созданных без указания артикула.
        for _ in range(20):
            candidate = f"AUTO-{uuid4().hex[:8].upper()}"
            if self.product_repository.get_by_article(candidate) is None:
                return candidate
        return f"AUTO-{uuid4().hex.upper()}"

    def create_product(self, payload: ProductCreatePayload, current_user: dict) -> dict:
        data = model_to_dict(payload)
        article = (data.get("product_article") or "").strip()
        if article:
            if self.product_repository.get_by_article(article) is not None:
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Товар с таким артикулом уже существует")
            data["product_article"] = article
        else:
            data["product_article"] = self._generate_unique_article()
        if data.get("product_cost_usd") is None:
            data["product_cost_usd"] = Decimal("0")
        row = self.product_repository.create({**data, "product_owner_user_id": current_user["user_id"]})
        log_audit_event(self.db, actor_user_id=current_user["user_id"], entity_type=ENTITY_TYPE_PRODUCT, entity_id=row.product_id, event_type=EVENT_TYPE_PRODUCT_CREATE, event_payload={"product_article": row.product_article})
        return serialize_product(row)

    def update_product(self, product_id: int, payload: ProductUpdatePayload, current_user: dict) -> dict:
        row = self.get_product_or_404(product_id)
        data = model_to_dict(payload)
        if not data:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нет данных для обновления")
        if "product_article" in data:
            existing = self.product_repository.get_by_article(data["product_article"])
            if existing is not None and existing.product_id != product_id:
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Товар с таким артикулом уже существует")
        row = self.product_repository.update(row, data)
        log_audit_event(self.db, actor_user_id=current_user["user_id"], entity_type=ENTITY_TYPE_PRODUCT, entity_id=row.product_id, event_type=EVENT_TYPE_PRODUCT_UPDATE, event_payload={"changed_fields": sorted(data.keys())})
        return serialize_product(row)

    def delete_product(self, product_id: int, current_user: dict) -> None:
        row = self.get_product_or_404(product_id)
        article = row.product_article
        self.product_repository.delete(row)
        log_audit_event(self.db, actor_user_id=current_user["user_id"], entity_type=ENTITY_TYPE_PRODUCT, entity_id=product_id, event_type=EVENT_TYPE_PRODUCT_DELETE, event_payload={"product_article": article})

    def import_products_from_xlsx(self, *, content: bytes, filename: str | None, current_user: dict) -> dict:
        if filename and not filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нужен файл формата .xlsx")

        parsed_rows = _iter_products_from_xlsx(content)
        job = _product_import_jobs.create(filename=filename, created_by_user_id=current_user["user_id"])
        _product_import_jobs.update(job["job_id"], total_rows=len(parsed_rows))
        session_factory = sessionmaker(bind=self.db.get_bind(), autocommit=False, autoflush=False, expire_on_commit=False)
        thread = Thread(
            target=_run_product_import_job,
            kwargs={
                "job_id": job["job_id"],
                "parsed_rows": parsed_rows,
                "current_user": current_user,
                "session_factory": session_factory,
            },
            daemon=True,
        )
        thread.start()
        return _product_import_jobs.get(job["job_id"])


def ensure_reference_data(db: Session) -> None:
    ReferenceDataService(db).ensure_seed_data()


def get_reference_data(db: Session) -> dict:
    return ReferenceDataService(db).get_reference_data()


def list_establishments(db: Session) -> dict:
    return ReferenceDataService(db).list_establishments()


def list_order_methods(db: Session) -> dict:
    return ReferenceDataService(db).list_order_methods()


def list_statuses(db: Session, *, status_type: str | None = None) -> dict:
    return ReferenceDataService(db).list_statuses(status_type=status_type)


def list_currencies(db: Session) -> dict:
    return ReferenceDataService(db).list_currencies()


def create_establishment(db: Session, payload: EstablishmentPayload, current_user: dict) -> dict:
    return ReferenceDataService(db).create_establishment(payload, current_user)


def create_order_method(db: Session, payload: OrderMethodPayload, current_user: dict) -> dict:
    return ReferenceDataService(db).create_order_method(payload, current_user)


def create_status(db: Session, payload: StatusPayload, current_user: dict) -> dict:
    return ReferenceDataService(db).create_status(payload, current_user)


def create_currency(db: Session, payload: CurrencyPayload, current_user: dict) -> dict:
    return ReferenceDataService(db).create_currency(payload, current_user)


def update_establishment(db: Session, establishment_id: int, payload: EstablishmentPayload) -> dict:
    return ReferenceDataService(db).update_establishment(establishment_id, payload)


def update_order_method(db: Session, order_method_id: int, payload: OrderMethodPayload) -> dict:
    return ReferenceDataService(db).update_order_method(order_method_id, payload)


def update_status(db: Session, status_id: int, payload: StatusPayload) -> dict:
    return ReferenceDataService(db).update_status(status_id, payload)


def update_currency(db: Session, currency_id: int, payload: CurrencyPayload) -> dict:
    return ReferenceDataService(db).update_currency(currency_id, payload)


def list_products(db: Session, *, search: str | None = None, page: int = 1, page_size: int = 20, sort_by: str = "product_id", sort_order: str = "desc") -> dict:
    return ReferenceDataService(db).list_products(search=search, page=page, page_size=page_size, sort_by=sort_by, sort_order=sort_order)


def create_product(db: Session, payload: ProductCreatePayload, current_user: dict) -> dict:
    return ReferenceDataService(db).create_product(payload, current_user)


def match_products_by_name(db: Session, names: list[str]) -> dict:
    return ReferenceDataService(db).match_products_by_name(names)


def update_product(db: Session, product_id: int, payload: ProductUpdatePayload, current_user: dict) -> dict:
    return ReferenceDataService(db).update_product(product_id, payload, current_user)


def delete_product(db: Session, product_id: int, current_user: dict) -> None:
    return ReferenceDataService(db).delete_product(product_id, current_user)


def import_products_from_xlsx(db: Session, *, content: bytes, filename: str | None, current_user: dict) -> dict:
    return ReferenceDataService(db).import_products_from_xlsx(content=content, filename=filename, current_user=current_user)


def get_product_import_job(job_id: str) -> dict:
    job = _product_import_jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Задача импорта не найдена")
    return job